from __future__ import annotations

from copy import copy
from io import BytesIO
import base64
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT_DIR / "users CACER.xlsx"
DEFAULT_OUTPUT = ROOT_DIR / "users CACER_dashboard.xlsx"
SCRIPT_DIR = Path(__file__).resolve().parent
ASSET_DIR_CANDIDATES = [SCRIPT_DIR / "assets", ROOT_DIR / "assets", SCRIPT_DIR, ROOT_DIR]
FAVICON_FILENAMES = ["favicon.png", "favicon(1).png"]
BACKGROUND_FILENAMES = ["background.jpg", "background.png"]
COMPANY_LOGO_FILENAMES = ["logo_RSE_1.png", "logo_RSE.png", "logo_rse.png"]
PALETTE = {
    "cyan": "#48C8D2", "cyan_dark": "#2388B5", "blue": "#2E78BD",
    "green": "#70C486", "green_dark": "#26A647", "ink": "#102331",
    "paper": "#F5FBFC", "muted": "#60717D",
}
USERS_SHEET = "Utenti"
DESCRIPTION_SHEET = "Description"

DESCRIPTION_COLUMNS = {
    "power_range": "A",
    "category": "D",
    "type": "F",
    "funding_scheme": "H",
    "supplier": "K",
    "tariff": "M",
    "load_profile_id": "Q",
    "building_archetype": "S",
    "hvac_type": "U",
}

BOOL_COLUMNS = {
    "flag",
    "consuming",
    "producing",
    "flag_DSM",
    "new_plant",
    "mv_cabinet",
    "heat_load",
    "flag_cacer",
    "dummy_user",
    "condominium",
}

INT_COLUMNS = {
    "number_type_id",
    "num",
    "pv",
    "tilt_angle",
    "azimuth",
    "wind",
    "battery",
    "th_comfort_heating",
    "th_comfort_cooling",
    "disbursement_month",
    "commissioning_month",
    "entry_month",
}

FLOAT_COLUMNS = {"grant_private", "grant_pnrr", "debt", "error"}
OPTIONAL_NUMERIC_COLUMNS = {
    "pv",
    "tilt_angle",
    "azimuth",
    "wind",
    "battery",
    "th_comfort_heating",
    "th_comfort_cooling",
}



def find_asset(filenames: list[str]) -> Path | None:
    for folder in ASSET_DIR_CANDIDATES:
        for filename in filenames:
            path = folder / filename
            if path.exists():
                return path
    return None


def image_to_base64(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def inject_dashboard_style(dark_mode: bool) -> None:
    background_b64 = image_to_base64(find_asset(BACKGROUND_FILENAMES))
    if dark_mode:
        theme = {
            "page_bg": "#07151D", "card_bg": "rgba(11, 31, 42, 0.88)",
            "card_bg_2": "rgba(16, 35, 49, 0.94)", "text": "#EEF8FA",
            "muted": "#B6CAD2", "border": "rgba(112, 196, 134, 0.24)",
            "input_bg": "rgba(255, 255, 255, 0.06)",
            "shadow": "0 20px 60px rgba(0, 0, 0, 0.32)",
            "overlay": "linear-gradient(135deg, rgba(7,21,29,0.94), rgba(7,21,29,0.78))",
        }
    else:
        theme = {
            "page_bg": PALETTE["paper"], "card_bg": "rgba(255, 255, 255, 0.86)",
            "card_bg_2": "rgba(255, 255, 255, 0.94)", "text": PALETTE["ink"],
            "muted": PALETTE["muted"], "border": "rgba(35, 136, 181, 0.16)",
            "input_bg": "rgba(255, 255, 255, 0.82)",
            "shadow": "0 20px 60px rgba(23, 86, 115, 0.16)",
            "overlay": "linear-gradient(135deg, rgba(245,251,252,0.93), rgba(245,251,252,0.78))",
        }

    background_css = ""
    if background_b64:
        background_css = f'''
        .stApp::before {{
            content: "";
            position: fixed;
            inset: 0;
            background-image: {theme["overlay"]}, url("data:image/jpg;base64,{background_b64}");
            background-size: cover;
            background-position: center;
            background-attachment: fixed;
            z-index: -2;
        }}
        '''

    st.markdown(f'''
        <style>
        :root {{
            --cacer-cyan: {PALETTE['cyan']}; --cacer-blue: {PALETTE['blue']};
            --cacer-green: {PALETTE['green']}; --cacer-green-dark: {PALETTE['green_dark']};
            --cacer-text: {theme['text']}; --cacer-muted: {theme['muted']};
            --cacer-card: {theme['card_bg']}; --cacer-card-strong: {theme['card_bg_2']};
            --cacer-border: {theme['border']}; --cacer-input: {theme['input_bg']};
            --cacer-shadow: {theme['shadow']};
        }}
        .stApp {{ background: {theme['page_bg']}; color: var(--cacer-text); }}
        {background_css}
        .block-container {{ padding-top: 1.6rem; max-width: 1420px; }}
        [data-testid="stSidebar"] > div:first-child {{
            background: linear-gradient(180deg, rgba(72,200,210,0.18), rgba(112,196,134,0.12)), var(--cacer-card-strong);
            border-right: 1px solid var(--cacer-border); backdrop-filter: blur(18px);
        }}
        .cacer-topbar {{ display: flex; align-items: center; justify-content: space-between; gap: 1rem;
            padding: 1rem 1.2rem; margin-bottom: 1.1rem; background: var(--cacer-card);
            border: 1px solid var(--cacer-border); border-radius: 24px; box-shadow: var(--cacer-shadow);
            backdrop-filter: blur(18px); }}
        .cacer-brand {{ display: flex; align-items: center; gap: 0.9rem; }}
        .cacer-project-icon {{ width: 58px; height: 58px; object-fit: contain; filter: drop-shadow(0 10px 18px rgba(35,136,181,0.24)); }}
        .cacer-eyebrow {{ margin: 0; font-size: 0.78rem; letter-spacing: 0.12em; text-transform: uppercase; color: var(--cacer-green-dark); font-weight: 800; }}
        .cacer-title {{ margin: 0; font-size: clamp(1.7rem, 3vw, 2.45rem); line-height: 1.05; color: var(--cacer-text); font-weight: 850; }}
        .cacer-subtitle {{ margin: 0.25rem 0 0; color: var(--cacer-muted); font-size: 0.98rem; }}
        .cacer-company-logo {{ max-width: 260px; width: 26vw; min-width: 160px; object-fit: contain; }}
        h1 {{ display: none; }} h2, h3 {{ color: var(--cacer-text); }}
        [data-testid="stForm"], [data-testid="stDataFrame"], .stAlert {{ background: var(--cacer-card);
            border: 1px solid var(--cacer-border); border-radius: 22px; box-shadow: var(--cacer-shadow); backdrop-filter: blur(18px); }}
        [data-testid="stForm"] {{ padding: 1.2rem; }}
        div[data-testid="stForm"] h3 {{ padding: 0.65rem 0.85rem; margin-top: 0.4rem; border-radius: 16px;
            background: linear-gradient(90deg, rgba(72,200,210,0.18), rgba(112,196,134,0.18)); border-left: 5px solid var(--cacer-cyan); }}
        .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {{ border: 0; border-radius: 16px;
            background: linear-gradient(135deg, var(--cacer-cyan), var(--cacer-green)); color: #ffffff; font-weight: 800;
            box-shadow: 0 12px 28px rgba(35,136,181,0.22); }}
        .stButton > button:hover, .stDownloadButton > button:hover, .stFormSubmitButton > button:hover {{ transform: translateY(-1px); filter: brightness(1.03); color: #ffffff; }}
        div[data-baseweb="input"] > div, div[data-baseweb="select"] > div, div[data-baseweb="textarea"] > div {{ background: var(--cacer-input); border-color: var(--cacer-border); border-radius: 14px; }}
        label, .stMarkdown, .stCaptionContainer, .stTextInput, .stNumberInput, .stSelectbox, .stCheckbox {{ color: var(--cacer-text) !important; }}
        hr {{ border-color: var(--cacer-border); }}
        </style>
    ''', unsafe_allow_html=True)


def render_header() -> None:
    project_icon_b64 = image_to_base64(find_asset(FAVICON_FILENAMES))
    company_logo_b64 = image_to_base64(find_asset(COMPANY_LOGO_FILENAMES))
    project_img = f'<img class="cacer-project-icon" src="data:image/png;base64,{project_icon_b64}" alt="RSE CoLabs">' if project_icon_b64 else ""
    company_img = f'<img class="cacer-company-logo" src="data:image/png;base64,{company_logo_b64}" alt="RSE">' if company_logo_b64 else ""
    st.markdown(f'''
        <div class="cacer-topbar">
            <div class="cacer-brand">
                {project_img}
                <div>
                    <p class="cacer-eyebrow">RSE CoLabs</p>
                    <p class="cacer-title">CACER users dashboard</p>
                    <p class="cacer-subtitle">Costruzione guidata del file <code>users CACER.xlsx</code></p>
                </div>
            </div>
            {company_img}
        </div>
    ''', unsafe_allow_html=True)

def as_clean_list(values: list[Any]) -> list[Any]:
    cleaned = []
    for value in values:
        if value is None or value == "":
            continue
        if value not in cleaned:
            cleaned.append(value)
    return cleaned


@st.cache_data(show_spinner=False)
def load_metadata(template_path: str) -> dict[str, Any]:
    workbook = load_workbook(template_path, data_only=True)
    if USERS_SHEET not in workbook.sheetnames or DESCRIPTION_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Il template deve contenere i fogli '{USERS_SHEET}' e '{DESCRIPTION_SHEET}'."
        )

    users_ws = workbook[USERS_SHEET]
    description_ws = workbook[DESCRIPTION_SHEET]
    headers = [
        users_ws.cell(row=1, column=col_idx).value
        for col_idx in range(1, users_ws.max_column + 1)
    ]
    headers = [header for header in headers if header]

    options = {}
    for field_name, column_letter in DESCRIPTION_COLUMNS.items():
        options[field_name] = as_clean_list(
            [
                description_ws[f"{column_letter}{row_idx}"].value
                for row_idx in range(2, description_ws.max_row + 1)
            ]
        )

    examples = []
    for row_idx in range(2, users_ws.max_row + 1):
        row = {
            header: users_ws.cell(row=row_idx, column=col_idx).value
            for col_idx, header in enumerate(headers, start=1)
        }
        if row.get("user_type"):
            examples.append(row)

    return {"headers": headers, "options": options, "examples": examples}


def empty_to_none(value: Any) -> Any:
    if value == "" or value is pd.NA:
        return None
    return value


def normalize_row(row: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    normalized = {}
    for header in headers:
        value = empty_to_none(row.get(header))
        if header in BOOL_COLUMNS:
            value = bool(value)
        elif header in INT_COLUMNS and value is not None:
            value = int(value)
        elif header in FLOAT_COLUMNS and value is not None:
            value = float(value)
        normalized[header] = value
    return normalized


def select_option(label: str, field: str, options: dict[str, list[Any]], value: Any, key: str) -> Any:
    choices = [""] + [str(option) for option in options.get(field, [])]
    current = "" if value in (None, "") else str(value)
    if current not in choices:
        choices.append(current)
    return st.selectbox(label, choices, index=choices.index(current), key=key) or None


def bool_input(label: str, value: Any, key: str) -> bool:
    return st.checkbox(label, value=bool(value), key=key)


def text_input(label: str, value: Any, key: str) -> str | None:
    return st.text_input(label, value="" if value is None else str(value), key=key).strip() or None


def numeric_input(
    label: str,
    value: Any,
    key: str,
    step: float = 1.0,
    optional: bool = False,
) -> int | float | None:
    if optional:
        enabled = st.checkbox(
            f"{label}: valorizza",
            value=value not in (None, ""),
            key=f"{key}_enabled",
        )
        if not enabled:
            return None

    default = 0.0 if value in (None, "") else float(value)
    result = st.number_input(label, value=default, step=step, key=key)
    return int(result) if float(result).is_integer() else result


def build_default_row(headers: list[str], examples: list[dict[str, Any]], selected_example: str) -> dict[str, Any]:
    if selected_example != "Nuova tipologia vuota":
        for example in examples:
            if example.get("user_type") == selected_example:
                return {header: example.get(header) for header in headers}

    row = {header: None for header in headers}
    row.update(
        {
            "number_type_id": len(st.session_state.user_rows) + 1,
            "flag": True,
            "num": 1,
            "type": "consumer",
            "consuming": True,
            "producing": False,
            "flag_DSM": False,
            "new_plant": True,
            "voltage": "BT",
            "mv_cabinet": False,
            "heat_load": False,
            "error": 0,
            "flag_cacer": True,
            "dummy_user": False,
            "condominium": False,
            "grant_private": 0,
            "grant_pnrr": 0,
            "debt": 0,
            "disbursement_month": 1,
            "commissioning_month": 1,
            "entry_month": 1,
            "exit_month": "end",
            "CP": "CP1",
            "location": "Milano",
        }
    )
    return row


def export_workbook(template_path: Path, rows: list[dict[str, Any]], headers: list[str]) -> bytes:
    workbook = load_workbook(template_path)
    users_ws = workbook[USERS_SHEET]

    for row_idx in range(2, users_ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            users_ws.cell(row=row_idx, column=col_idx).value = None

    style_source_row = 2
    for row_offset, row in enumerate(rows, start=2):
        clean_row = normalize_row(row, headers)
        for col_idx, header in enumerate(headers, start=1):
            target = users_ws.cell(row=row_offset, column=col_idx)
            source = users_ws.cell(row=style_source_row, column=col_idx)
            if row_offset != style_source_row:
                target._style = copy(source._style)
                if source.has_style:
                    target.font = copy(source.font)
                    target.fill = copy(source.fill)
                    target.border = copy(source.border)
                    target.alignment = copy(source.alignment)
                    target.number_format = source.number_format
                    target.protection = copy(source.protection)
            target.value = clean_row.get(header)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def add_or_replace_row(row: dict[str, Any]) -> None:
    user_type = row.get("user_type")
    existing_idx = next(
        (
            idx
            for idx, existing_row in enumerate(st.session_state.user_rows)
            if existing_row.get("user_type") == user_type
        ),
        None,
    )
    if existing_idx is None:
        st.session_state.user_rows.append(row)
    else:
        st.session_state.user_rows[existing_idx] = row


def render_row_form(base_row: dict[str, Any], metadata: dict[str, Any]) -> dict[str, Any] | None:
    options = metadata["options"]
    row = dict(base_row)

    with st.form("user_type_form", clear_on_submit=False):
        st.subheader("1. Identificazione")
        c1, c2, c3 = st.columns(3)
        with c1:
            row["number_type_id"] = numeric_input("number_type_id", row.get("number_type_id"), "number_type_id")
            row["flag"] = bool_input("flag", row.get("flag", True), "flag")
            row["user_type"] = text_input("user_type", row.get("user_type"), "user_type")
        with c2:
            row["num"] = numeric_input("num", row.get("num", 1), "num")
            row["denomination"] = text_input("denomination", row.get("denomination"), "denomination")
            row["stakeholder"] = text_input("stakeholder", row.get("stakeholder"), "stakeholder")
        with c3:
            row["CP"] = text_input("CP", row.get("CP"), "CP")
            row["location"] = text_input("location", row.get("location"), "location")
            row["category"] = select_option("category", "category", options, row.get("category"), "category")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            row["type"] = select_option("type", "type", options, row.get("type"), "type")
        with c2:
            row["consuming"] = bool_input("consuming", row.get("consuming"), "consuming")
        with c3:
            row["producing"] = bool_input("producing", row.get("producing"), "producing")
        with c4:
            row["flag_cacer"] = bool_input("flag_cacer", row.get("flag_cacer", True), "flag_cacer")

        st.subheader("2. Consumi e connessione")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            row["power_range"] = select_option(
                "power_range", "power_range", options, row.get("power_range"), "power_range"
            )
        with c2:
            row["supplier"] = select_option("supplier", "supplier", options, row.get("supplier"), "supplier")
        with c3:
            row["tariff"] = select_option("tariff", "tariff", options, row.get("tariff"), "tariff")
        with c4:
            row["load_profile_id"] = select_option(
                "load_profile_id",
                "load_profile_id",
                options,
                row.get("load_profile_id"),
                "load_profile_id",
            )

        c1, c2, c3 = st.columns(3)
        with c1:
            row["flag_DSM"] = bool_input("flag_DSM", row.get("flag_DSM"), "flag_DSM")
        with c2:
            row["voltage"] = st.selectbox(
                "voltage",
                ["BT", "MT"],
                index=["BT", "MT"].index(row.get("voltage", "BT") if row.get("voltage") in ["BT", "MT"] else "BT"),
                key="voltage",
            )
        with c3:
            row["mv_cabinet"] = bool_input("mv_cabinet", row.get("mv_cabinet"), "mv_cabinet")

        st.subheader("3. Generazione e accumulo")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            row["pv"] = numeric_input("pv", row.get("pv"), "pv", optional=True)
        with c2:
            row["tilt_angle"] = numeric_input(
                "tilt_angle", row.get("tilt_angle"), "tilt_angle", optional=True
            )
        with c3:
            row["azimuth"] = numeric_input("azimuth", row.get("azimuth"), "azimuth", optional=True)
        with c4:
            row["pv_mounting"] = st.selectbox(
                "pv_mounting",
                ["", "roof", "ground"],
                index=["", "roof", "ground"].index(row.get("pv_mounting") or ""),
                key="pv_mounting",
            ) or None

        c1, c2, c3 = st.columns(3)
        with c1:
            row["wind"] = numeric_input("wind", row.get("wind"), "wind", optional=True)
        with c2:
            row["battery"] = numeric_input("battery", row.get("battery"), "battery", optional=True)
        with c3:
            row["new_plant"] = bool_input("new_plant", row.get("new_plant", True), "new_plant")

        st.subheader("4. Termico")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            row["building_archetype"] = select_option(
                "building_archetype",
                "building_archetype",
                options,
                row.get("building_archetype"),
                "building_archetype",
            )
        with c2:
            row["hvac_type"] = select_option("hvac_type", "hvac_type", options, row.get("hvac_type"), "hvac_type")
        with c3:
            row["th_comfort_heating"] = numeric_input(
                "th_comfort_heating",
                row.get("th_comfort_heating"),
                "th_comfort_heating",
                optional=True,
            )
        with c4:
            row["th_comfort_cooling"] = numeric_input(
                "th_comfort_cooling",
                row.get("th_comfort_cooling"),
                "th_comfort_cooling",
                optional=True,
            )

        c1, c2 = st.columns(2)
        with c1:
            row["heat_load"] = bool_input("heat_load", row.get("heat_load"), "heat_load")
        with c2:
            row["condominium"] = bool_input("condominium", row.get("condominium"), "condominium")

        st.subheader("5. Finanza e calendario")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            row["funding_scheme"] = select_option(
                "funding_scheme",
                "funding_scheme",
                options,
                row.get("funding_scheme"),
                "funding_scheme",
            )
        with c2:
            row["dummy_user"] = bool_input("dummy_user", row.get("dummy_user"), "dummy_user")
        with c3:
            row["grant_private"] = numeric_input(
                "grant_private", row.get("grant_private", 0), "grant_private", step=0.01
            )
        with c4:
            row["grant_pnrr"] = numeric_input("grant_pnrr", row.get("grant_pnrr", 0), "grant_pnrr", step=0.01)

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            row["debt"] = numeric_input("debt", row.get("debt", 0), "debt", step=0.01)
        with c2:
            row["disbursement_month"] = numeric_input(
                "disbursement_month", row.get("disbursement_month", 1), "disbursement_month"
            )
        with c3:
            row["commissioning_month"] = numeric_input(
                "commissioning_month", row.get("commissioning_month", 1), "commissioning_month"
            )
        with c4:
            row["entry_month"] = numeric_input("entry_month", row.get("entry_month", 1), "entry_month")

        c1, c2 = st.columns(2)
        with c1:
            row["exit_month"] = text_input("exit_month", row.get("exit_month", "end"), "exit_month")
        with c2:
            row["error"] = numeric_input("error", row.get("error", 0), "error", step=0.01)

        submitted = st.form_submit_button("Aggiungi o aggiorna tipologia", use_container_width=True)

    if submitted:
        if not row.get("user_type"):
            st.error("Inserisci un valore per user_type.")
            return None
        if len(str(row["user_type"])) > 31:
            st.error("user_type deve avere al massimo 31 caratteri, altrimenti Excel non potra usarlo come nome foglio.")
            return None
        funding_total = sum(float(row.get(field) or 0) for field in ["grant_private", "grant_pnrr", "debt"])
        if funding_total > 1:
            st.error("La somma grant_private + grant_pnrr + debt deve essere minore o uguale a 1.")
            return None
        return row
    return None


def main() -> None:
    favicon_path = find_asset(FAVICON_FILENAMES)
    st.set_page_config(
        page_title="CACER users dashboard",
        page_icon=str(favicon_path) if favicon_path else "⚡",
        layout="wide",
    )

    if "dark_mode" not in st.session_state:
        st.session_state.dark_mode = False
    dark_mode = st.sidebar.toggle(
        "🌙 Modalità notte",
        value=st.session_state.dark_mode,
        help="Passa dalla modalità giorno alla modalità notte senza modificare i dati inseriti.",
    )
    st.session_state.dark_mode = dark_mode

    inject_dashboard_style(dark_mode)
    render_header()

    if "user_rows" not in st.session_state:
        st.session_state.user_rows = []

    template_path = st.sidebar.text_input("Template Excel", value=str(DEFAULT_TEMPLATE))
    output_path = st.sidebar.text_input("File output", value=str(DEFAULT_OUTPUT))
    template = Path(template_path)

    try:
        metadata = load_metadata(str(template))
    except Exception as exc:
        st.error(f"Impossibile leggere il template: {exc}")
        return

    examples = metadata["examples"]
    example_names = ["Nuova tipologia vuota"] + [str(row["user_type"]) for row in examples]
    selected_example = st.sidebar.selectbox("Base di partenza", example_names)

    if st.sidebar.button("Importa righe attive dal template"):
        st.session_state.user_rows = [
            {header: row.get(header) for header in metadata["headers"]}
            for row in examples
            if row.get("flag") is True
        ]
        st.sidebar.success(f"Importate {len(st.session_state.user_rows)} righe attive.")

    if st.sidebar.button("Svuota righe in costruzione"):
        st.session_state.user_rows = []

    base_row = build_default_row(metadata["headers"], examples, selected_example)
    submitted_row = render_row_form(base_row, metadata)
    if submitted_row is not None:
        add_or_replace_row(submitted_row)
        st.success(f"Tipologia `{submitted_row['user_type']}` aggiunta/aggiornata.")

    st.divider()
    st.subheader("6. Anteprima ed export")

    rows = st.session_state.user_rows
    if rows:
        preview = pd.DataFrame(rows, columns=metadata["headers"])
        st.dataframe(preview, use_container_width=True, hide_index=True)

        user_type_to_remove = st.selectbox(
            "Rimuovi una tipologia dall'anteprima",
            [""] + [str(row.get("user_type")) for row in rows],
        )
        if st.button("Rimuovi selezionata", disabled=not user_type_to_remove):
            st.session_state.user_rows = [
                row for row in rows if str(row.get("user_type")) != user_type_to_remove
            ]
            st.rerun()

        workbook_bytes = export_workbook(template, rows, metadata["headers"])
        st.download_button(
            "Scarica Excel generato",
            data=workbook_bytes,
            file_name=Path(output_path).name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        if st.button("Salva Excel nel repository", use_container_width=True):
            destination = Path(output_path)
            destination.write_bytes(workbook_bytes)
            st.success(f"File salvato in: {destination}")
    else:
        st.info("Aggiungi almeno una tipologia per generare il file Excel.")


if __name__ == "__main__":
    main()
